#!/usr/bin/env python3
"""
Genera Inventario_Equipos_DMZ.xlsx y su versión TXT.
"""
import sys, json, glob, math, re
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tabulate import tabulate

# Utilidades

def parse_size(size_str: str) -> float:
    m = re.match(r"([\d.]+)\s*(GB|MB|TB)", size_str or "")
    if not m:
        return 0.0
    val, unit = float(m[1]), m[2]
    return val / 1024 if unit == "MB" else val * 1024 if unit == "TB" else val

def pick(d: dict, *keys):
    for k in keys:
        if isinstance(k, (list, tuple)):
            for sub in k:
                if sub in d:
                    return d[sub]
        elif k in d:
            return d[k]
    return None

def fila_equipo(facts: dict, host_inv: str) -> dict:
    ipdata = pick(facts, "ansible_default_ipv4", "default_ipv4") or {}
    ip = ipdata.get("address") or host_inv
    hostname = pick(facts, "ansible_hostname", "hostname", "fqdn") or host_inv

    distro = pick(facts, "ansible_distribution", "distribution") or ""
    d_ver = pick(facts, "ansible_distribution_version", "distribution_version") or ""
    so = f"{distro} {d_ver}".strip()

    kernel = pick(facts, "ansible_kernel", "kernel") or ""
    arch = pick(facts, "ansible_architecture", "architecture", "machine") or ""

    cpu_lst = pick(facts, "ansible_processor", "processor") or []
    cpu = cpu_lst[2] if len(cpu_lst) > 2 else cpu_lst[-1] if cpu_lst else ""

    mem_total_mb = pick(facts, "ansible_memtotal_mb", "memtotal_mb") or 0
    mem_free_mb = pick(facts, "ansible_memfree_mb", "memfree_mb", ["ansible_memory_mb", "memory_mb"]) or 0
    if isinstance(mem_free_mb, dict):
        mem_free_mb = mem_free_mb.get("real", {}).get("free", 0)
    mem_total_gb = math.ceil(mem_total_mb / 1024)
    mem_free_gb = math.ceil(mem_free_mb / 1024)
    mem_used_gb = mem_total_gb - mem_free_gb

    devs = pick(facts, "ansible_devices", "devices") or {}
    disk_total_gb = sum(parse_size(d.get("size", "0 GB")) for n, d in devs.items() if n.startswith(("sd", "nvme")))

    tipo = "Virtual" if pick(facts, "ansible_virtualization_role", "virtualization_role") == "guest" else "Física"
    puertos_txt = ", ".join(map(str, facts.get("listening_ports", [])))

    db_cols = {
        "MySQL (3306)": "Activo" if facts.get("mysql") else "Inactivo",
        "PostgreSQL (5432)": "Activo" if facts.get("postgresql") else "Inactivo",
        "SQLServer (1433)": "Activo" if facts.get("sqlserver") else "Inactivo",
        "Oracle (1521)": "Activo" if facts.get("oracle") else "Inactivo",
        "MongoDB (27017)": "Activo" if facts.get("mongodb") else "Inactivo",
    }

    return {
        "FechaHora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "IP": ip,
        "Host": hostname,
        "SO": so,
        "CPU": cpu,
        "RAMTot": mem_total_gb,
        "RAMUsed": mem_used_gb,
        "DSKT": disk_total_gb,
        "Tipo": tipo,
        "Puertos": puertos_txt,
        **db_cols,
        "Kernel": kernel,
        "Arch": arch,
        "RAMFree": mem_free_gb,
    }

def extraer_usuarios_grupos(datos_json: dict):
    usuarios_brutos = datos_json.get("usuarios", [])
    usuarios = []
    grupos = []
    modo = None

    for linea in usuarios_brutos:
        if "Usuarios del sistema" in linea:
            modo = "usuarios"
            continue
        elif "Grupos del sistema" in linea:
            modo = "grupos"
            continue

        if modo == "usuarios":
            match = re.match(r"([\w\-]+)\s+\(UID:\s*(\d+),\s*GID:\s*(\d+),\s*Shell:\s*(.*?)\)", linea)
            if match:
                nombre, uid, gid, shell = match.groups()
                login = "Sí" if shell not in ["/usr/sbin/nologin", "/bin/false", "nologin"] else "No"
                usuarios.append({"Usuario": nombre, "UID": uid, "GID": gid, "Login": login})
        elif modo == "grupos":
            partes = linea.split(":")
            if len(partes) >= 4:
                grupo = partes[0]
                miembros = partes[3].split(",") if partes[3] else []
                grupos.append({"Grupo": grupo, "Miembros": ", ".join([m.strip() for m in miembros if m])})

    return usuarios, grupos

def main():
    if len(sys.argv) < 3:
        sys.exit("Uso: generar_excel.py <salida.xlsx> <glob_json>")

    salida_excel, patron = sys.argv[1], sys.argv[2]
    filas_excel = []
    contenido_txt = []

    for path in glob.glob(patron):
        with open(path, "r") as f:
            info = json.load(f)
        facts = info.get("ansible_facts", info)

        fila = fila_equipo(facts, info.get("inventory_hostname", "desconocido"))
        filas_excel.append(fila)

        usuarios, grupos = extraer_usuarios_grupos(info)

        contenido_txt.append(f"=== {fila['Host']} ({fila['IP']}) ===\n")
        contenido_txt.append(tabulate(usuarios, headers="keys", tablefmt="pretty", showindex=False))
        contenido_txt.append("\n")
        contenido_txt.append(tabulate(grupos, headers="keys", tablefmt="pretty", showindex=False))
        contenido_txt.append("\n\n")

    df = pd.DataFrame(filas_excel).sort_values("IP")
    with pd.ExcelWriter(salida_excel, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Inventario DMZ", index=False)
        hoja = writer.sheets["Inventario DMZ"]
        for i, col in enumerate(df.columns, 1):
            ancho = max(len(str(x)) for x in df[col].astype(str).tolist() + [col]) + 2
            hoja.column_dimensions[get_column_letter(i)].width = ancho
        n_filas, n_cols = df.shape
        ultima_col = get_column_letter(n_cols)
        tabla = Table(displayName="InventarioDMZ", ref=f"A1:{ultima_col}{n_filas+1}")
        tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        hoja.add_table(tabla)

    txt_path = Path(salida_excel).with_suffix("_usuarios.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(contenido_txt))

    print(f"✅ Excel generado → {salida_excel}")
    print(f"✅ TXT generado   → {txt_path}")

if __name__ == "__main__":
    main()
